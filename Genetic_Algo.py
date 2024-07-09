import random
from enum import Enum
from functools import lru_cache
from math import ceil
from statistics import mean
from typing import NamedTuple
from openpyxl import Workbook
import numpy as np
from deap import algorithms
from deap import base
from deap import creator
from deap import tools
import tkinter as tk
from tabulate import tabulate
import calendar
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

import warnings
warnings.simplefilter("ignore", UserWarning)
# Suppress the warning for redefining the NamedTuple class
warnings.filterwarnings("ignore", category=UserWarning, module="deap.creator")



# Add the missing Faculty class definition here
class Faculty:
    def __init__(self, name, subject_expertise):
        self.name = name
        self.subject_expertise = subject_expertise

SUBJECTS = []

class Class(NamedTuple):
    subject: 'Subject'
    type: int  # Use integers 0 for LECTURE, 1 for PRACTICAL
    faculty: Faculty

    def __repr__(self):
        return self.name

class Subject:
    def __init__(self, name, numbers, faculty):
        self.name = name
        self.numbers = numbers
        self.faculty = faculty

    def __repr__(self):
        return self.name

class SubjectType(Enum):
    LECTURE = 0
    PRACTICAL = 1

class WeekDay(Enum):
    MON = 0
    TUE = 1
    WED = 2
    THU = 3
    FRI = 4
    SAT = 5
    SUN = 6

C1 = 0.5  # uniformity
C2 = 1  # tightness
C3 = 1  # suitability
C4 = 1  # sleep
C5 = 1  # day_grouping
C6 = 1  # week_grouping

CLASSES_PER_DAY = 6
WORKING_DAYS = 5

PLACES = CLASSES_PER_DAY * WORKING_DAYS

# Define faculty members
FACULTY_MEMBERS = [
    Faculty(name='Mrs. T G Deokar', subject_expertise=['SPC']),             #0
    Faculty(name='Mrs. Deokar', subject_expertise=['EDAV', 'SS']),          #1
    Faculty(name='Mr. Kamble', subject_expertise=['ML', 'AML']),            #2
    Faculty(name='Mr. Kadam', subject_expertise=['Database']),              #3
    Faculty(name='Mr. Vadagave', subject_expertise=['Java', 'Database']),   #4
    Faculty(name='Mrs. Chavan', subject_expertise=['LA']),                  #5
    Faculty(name='Ms. A Patil', subject_expertise=['DMS']),                 #6
    Faculty(name='Mr. G V Patil', subject_expertise=['DS']),                #7
    Faculty(name='Mr. S Patil', subject_expertise=['FN']),                  #8
    Faculty(name='Mr. Metkari', subject_expertise=['PP']),
    Faculty(name='Mr. ABC', subject_expertise=['AI']),                 #10

]

# Define subjects for the year 1
YEAR_1_SUBJECTS = [
    Subject(name='SPC', numbers={0: 5}, faculty=FACULTY_MEMBERS[0]),
    Subject(name='EDAV', numbers={1: 5}, faculty=FACULTY_MEMBERS[1]),
    Subject(name='ML', numbers={0: 5}, faculty=FACULTY_MEMBERS[2]),
    Subject(name='Database', numbers={0: 5}, faculty=FACULTY_MEMBERS[3]),
    Subject(name='Java', numbers={0: 5}, faculty=FACULTY_MEMBERS[4]),
    Subject(name='AI', numbers={0: 5}, faculty=FACULTY_MEMBERS[10]),
    # Add more subjects for the first year as needed
]

# Define subjects for the year 2
YEAR_2_SUBJECTS = [
    Subject(name='LA', numbers={0: 5}, faculty=FACULTY_MEMBERS[5]),
    Subject(name='DMS', numbers={1: 5}, faculty=FACULTY_MEMBERS[6]),
    Subject(name='DS', numbers={0: 5}, faculty=FACULTY_MEMBERS[7]),
    Subject(name='FN', numbers={0: 5}, faculty=FACULTY_MEMBERS[8]),
    Subject(name='PP', numbers={0: 5}, faculty=FACULTY_MEMBERS[9]),
    Subject(name='AI', numbers={0: 5}, faculty=FACULTY_MEMBERS[10]),
    # Add more subjects for the second year as needed
]

# Add subjects for both years to the SUBJECTS list
SUBJECTS.extend(YEAR_1_SUBJECTS + YEAR_2_SUBJECTS)

# Create instances of the Class class for each year
year_1_lectures = [Class(subject, 0, subject.faculty) for subject in YEAR_1_SUBJECTS for _ in
                   range(subject.numbers.get(0, 0))]
year_1_practicals = [Class(subject, 1, subject.faculty) for subject in YEAR_1_SUBJECTS for _ in
                     range(subject.numbers.get(1, 0))]

year_2_lectures = [Class(subject, 0, subject.faculty) for subject in YEAR_2_SUBJECTS for _ in
                   range(subject.numbers.get(0, 0))]
year_2_practicals = [Class(subject, 1, subject.faculty) for subject in YEAR_2_SUBJECTS for _ in
                     range(subject.numbers.get(1, 0))]

# Combine classes for both years
classes = year_1_lectures + year_1_practicals + year_2_lectures + year_2_practicals

toolbox = base.Toolbox()


SUITABLE_TIME = {
    WeekDay.MON: {1: [], 2: [], 3: [], 4: [], 5: [], 6: []},
    WeekDay.TUE: {1: [], 2: [], 3: [], 4: [], 5: [], 6: []},
    WeekDay.WED: {1: [], 2: [], 3: [], 4: [], 5: [], 6: []},
    WeekDay.THU: {1: [], 2: [], 3: [], 4: [], 5: [], 6: []},
    WeekDay.FRI: {1: [], 2: [], 3: [], 4: [], 5: [], 6: []}
}

# for better performance
def diff(x):
    return [b - a for a, b in zip(x[:-1], x[1:])]

@lru_cache(maxsize=None)
def calc_uniformity(num_by_day):
    n = sum(num_by_day)
    p = WORKING_DAYS
    mean_num = n / p

    # for 14: 3 3 3 3 2
    min_dev = (p - n % p) * (mean_num - n // p) + \
              (n % p) * (n // p + 1 - mean_num)
    # stack all classes one after other i.e. for 14: 4 4 4 2 0
    max_dev = ((n // CLASSES_PER_DAY) * (CLASSES_PER_DAY - mean_num) +
               abs(n % CLASSES_PER_DAY - mean_num) +
               (p - ceil(n / CLASSES_PER_DAY)) * mean_num)
    dev = sum(abs(c - mean_num) for c in num_by_day)
    return 1 - (dev - min_dev) / (max_dev - min_dev)

def first_non_none(lst):
    return next(i for i, class_ in enumerate(lst) if class_ is not None)

def construct_schedule(ordering, mapping):
    classes_num = len(mapping)
    subjects_ordering = [mapping[o] if o < classes_num else None for o in ordering]
    return [subjects_ordering[i:i + CLASSES_PER_DAY] for i in range(0, len(subjects_ordering), CLASSES_PER_DAY)]

def eval_ordering(ordering, classes, mapping):
    classes_num = len(mapping)
    schedule = construct_schedule(ordering, mapping)
    num_by_day = tuple(sum(s is not None for s in day) for day in schedule)

    uniformity = calc_uniformity(num_by_day)

    tightness = 1 - sum(
        sum(class_ is None for class_ in day[first_non_none(day) + 1:first_non_none(reversed(day))])
        if any(class_ is not None for class_ in day) else 0
        for day in schedule
    ) / (PLACES - classes_num)

    suitability = sum(
        class_.subject.name in SUITABLE_TIME[week_day][class_num0 + 1]
        for day, week_day in zip(schedule, WeekDay)
        for class_num0, class_ in enumerate(day)
        if class_ is not None
    ) / classes_num
    sleep = sum(day[0] is None for day in schedule) / WORKING_DAYS

    unique_indexes = [
        [[i for i, class_ in enumerate(day) if class_ == unique_class]
         for unique_class in set(day).difference([None])]
        for day in schedule
    ]
    day_grouping = 1 - mean(
        sum(d - 1 for d in diff(class_indexes)) / (day_num - len(class_indexes))
        if len(day) > 1 else 0.
        for day, day_num in zip(unique_indexes, num_by_day) for class_indexes in day
    )

    act_tot_num = [(sum(class_ in day for day in schedule), class_.subject.numbers[class_.type]) for class_ in classes]
    act_min_max = [(act_num, ceil(tot_num / CLASSES_PER_DAY), min(tot_num, WORKING_DAYS))
                   for act_num, tot_num in act_tot_num]
    week_grouping = 1 - mean(
        ((act_num - min_num) / (max_num - min_num))
        if max_num - min_num != 0 else 1.
        for act_num, min_num, max_num in act_min_max
    )

    # Check faculty and subject constraints
    faculty_overlaps = any(len(set(day[i].faculty.name for i in range(len(day)) if day[i] is not None)) > 1
                           for day in schedule)
    subject_overlaps = any(len(set(day[i].subject.name for i in range(len(day)) if day[i] is not None)) > 1
                            for day in schedule)
    if faculty_overlaps or subject_overlaps:
        return (0.0,)  # If faculty or subject overlaps, the timetable is invalid

    return (C1 * uniformity + C2 * tightness + C3 * suitability + C4 * sleep + C5 * day_grouping + C6 * week_grouping,)


# ...
def create_toolbox(classes, mapping):
    toolbox = base.Toolbox()
    creator.create("FitnessMulti", base.Fitness, weights=(1.0,))
    creator.create("Individual", list, fitness=creator.FitnessMulti)

    toolbox.register('indices', random.sample, range(PLACES), PLACES)
    toolbox.register('individual', tools.initIterate, creator.Individual, toolbox.indices)
    toolbox.register('population', tools.initRepeat, list, toolbox.individual)

    # Pass the mapping to eval_ordering
    toolbox.register('evaluate', lambda ordering: eval_ordering(ordering, classes, mapping))
    toolbox.register('mate', tools.cxOrdered)
    toolbox.register('mutate', tools.mutShuffleIndexes, indpb=0.01)
    toolbox.register('select', tools.selTournament, tournsize=3)

    return toolbox


def write_timetable_to_excel(filename, schedule):
    wb = Workbook()
    ws = wb.active

    # Write headers
    headers = ["", "Class 1", "Class 2", "Class 3", "Class 4", "Class 5", "Class 6"]
    ws.append(headers)

    # Write schedule data
    for day_num, day in enumerate(schedule):
        row_data = [f"Day {day_num + 1}"] + [f"{class_.subject.name} - {class_.faculty.name}" if class_ is not None else "No class" for class_ in day]
        ws.append(row_data)

    # Save the workbook to the specified filename
    wb.save(filename)


def generate_timetable(classes, mapping, toolbox):
    population = toolbox.population(n=300)
    hof = tools.HallOfFame(1)
    stats = tools.Statistics(lambda ind: ind.fitness.values)
    stats.register('avg', np.mean)
    stats.register('std', np.std)
    stats.register('min', np.min)
    stats.register('max', np.max)

    algorithms.eaSimple(
        population, toolbox,
        ngen=300,
        cxpb=0.5,
        mutpb=0.2,
        stats=stats,
        verbose=False,
        halloffame=hof,
    )

    best_schedule = construct_schedule(hof.items[0], mapping)
    return best_schedule  # Return the generated timetable instead of the function



def print_timetable(schedule):
    table_data = []
    # Get the names of the days of the week
    days_of_week = list(calendar.day_name)
    for day_num, day in enumerate(schedule):
        row = [days_of_week[day_num]]
        for class_num, class_ in enumerate(day):
            if class_ is not None:
                row.append(f"{class_.subject.name} - {class_.faculty.name}")
            else:
                row.append("No class")
        table_data.append(row)

    headers = [""] + [f"Class {class_num + 1}" for class_num in range(CLASSES_PER_DAY)]
    print(tabulate(table_data, headers=headers))

def excel_to_customized_image(excel_file_path, output_image_path):
    # Load the Excel workbook
    wb = load_workbook(excel_file_path)
    sheet = wb.active

    # Get dimensions of the sheet
    max_row = sheet.max_row
    max_col = sheet.max_column

    # Create a new image (size depends on your template)
    image = Image.new("RGB", (1240, 620), "#2C4072")
    draw = ImageDraw.Draw(image)

    # Set font and other parameters
    font = ImageFont.truetype("arial.ttf", 19)

    # Define cell size and starting position
    cell_width = 165
    cell_height = 45
    x_start = 50
    y_start = 10

    # Iterate through each cell and draw it on the image
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell_value = sheet.cell(row=row, column=col).value

            # Skip drawing cells with specific values
            if cell_value in ["None", None]:  # Add values to skip here
                continue

            # Remove suffix for certain cells
            if "-" in str(cell_value):
                # Split the cell value based on "-"
                parts = str(cell_value).split("-", 1)
                # Take the part before "-" and remove leading/trailing whitespace
                cell_value = parts[0].strip()

            x = x_start + (col - 1) * cell_width
            y = y_start + (row - 1) * cell_height
            draw.rectangle([x, y, x + cell_width, y + cell_height], outline="black")
            draw.text((x + 5, y + 5), str(cell_value), fill="white", font=font)

    # Define cell size and starting position for the second table
    table2_x_start = 230
    table2_y_start = 295
    cell_width1 = 300

    # Manually set cell values for the second table
    table2_values = [
        ["Course Name", "Abbrevation", "Faculty Name"],
        ["Linear Algebra", "LA", "Ms. Chavan"],
        ["Python Programming", "PP", "Mr. Metkari"],
        ["Data Structures", "DS", "Mr. G.V. Patil"],
        ["Database Management Systems", "DMS", "Ms. A. Patil"],
        ["Fundamental of Networking", "FN", "Mr. S. Patil"]
    ]

    # Iterate through each cell and draw it on the image for the second table
    for row in range(1, 7):
        for col in range(1, 4):
            cell_value = table2_values[row - 1][col - 1]

            x = table2_x_start + (col - 1) * cell_width1
            y = table2_y_start + (row - 1) * cell_height
            draw.rectangle([x, y, x + cell_width1, y + cell_height], outline="black")
            draw.text((x + 5, y + 5), str(cell_value), fill="white", font=font)

    # Save the image
    image.save(output_image_path)

def excel_to_customized_image2(excel_file_path, output_image_path):
    # Load the Excel workbook
    wb = load_workbook(excel_file_path)
    sheet = wb.active

    # Get dimensions of the sheet
    max_row = sheet.max_row
    max_col = sheet.max_column

    # Create a new image (size depends on your template)
    image = Image.new("RGB", (1240, 620), "#2C4072")
    draw = ImageDraw.Draw(image)

    # Set font and other parameters
    font = ImageFont.truetype("arial.ttf", 19)

    # Define cell size and starting position
    cell_width = 165
    cell_height = 45
    x_start = 50
    y_start = 10

    # Iterate through each cell and draw it on the image
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell_value = sheet.cell(row=row, column=col).value

            # Skip drawing cells with specific values
            if cell_value in ["None", None]:  # Add values to skip here
                continue

            # Remove suffix for certain cells
            if "-" in str(cell_value):
                # Split the cell value based on "-"
                parts = str(cell_value).split("-", 1)
                # Take the part before "-" and remove leading/trailing whitespace
                cell_value = parts[0].strip()

            x = x_start + (col - 1) * cell_width
            y = y_start + (row - 1) * cell_height
            draw.rectangle([x, y, x + cell_width, y + cell_height], outline="black")
            draw.text((x + 5, y + 5), str(cell_value), fill="white", font=font)

    # Define cell size and starting position for the second table
    table2_x_start = 230
    table2_y_start = 295
    cell_width1 = 300

    # Manually set cell values for the second table
    table2_values = [
        ["Course Name", "Abbrevation", "Faculty Name"],
        ["Exploratory Data Analysis", "EDA", "Mrs. Deokar"],
        ["Java Programming", "JP", "Mr. Vadagave"],
        ["System Programming and Compilers", "SPC", "Mrs. T.G. Deokar"],
        ["Machine Learning", "ML", "Mr. Kamble"],
        ["Database", "DB", "Mr. Kadam"]
    ]

    # Iterate through each cell and draw it on the image for the second table
    for row in range(1, 7):
        for col in range(1, 4):
            cell_value = table2_values[row - 1][col - 1]

            x = table2_x_start + (col - 1) * cell_width1
            y = table2_y_start + (row - 1) * cell_height
            draw.rectangle([x, y, x + cell_width1, y + cell_height], outline="black")
            draw.text((x + 5, y + 5), str(cell_value), fill="white", font=font)

    # Save the image
    image.save(output_image_path)

# Modify your main function to call write_timetable_to_excel
def main():
    divisions = 2
    excel_to_customized_image(r".\2nd_Year_Division_1.xlsx", r".\2nd_Year_Division_1.png")
    excel_to_customized_image(r".\2nd_Year_Division_2.xlsx", r".\2nd_Year_Division_2.png")
    excel_to_customized_image2(r".\3rd_Year_Division_1.xlsx", r".\3rd_Year_Division_1.png")
    excel_to_customized_image2(r".\3rd_Year_Division_2.xlsx", r".\3rd_Year_Division_2.png")
    for schedule_num in range(divisions):
        print(f"3rd Year - Division {schedule_num + 1}:")
        year_1_lectures = [Class(subject, 0, subject.faculty) for subject in YEAR_1_SUBJECTS for _ in range(subject.numbers.get(0, 0))]
        year_1_practicals = [Class(subject, 1, subject.faculty) for subject in YEAR_1_SUBJECTS for _ in range(subject.numbers.get(1, 0))]
        classes = year_1_lectures + year_1_practicals
        mapping = {i: class_ for i, class_ in enumerate(classes)}
        toolbox = create_toolbox(classes, mapping)
        schedule = generate_timetable(classes, mapping, toolbox)  # Correctly call the function
        print_timetable(schedule)
        write_timetable_to_excel(f"3rd_Year_Division_{schedule_num + 1}.xlsx", schedule)

    for schedule_num in range(divisions):
        print(f"2nd Year - Division {schedule_num + 1}:")
        year_2_lectures = [Class(subject, 0, subject.faculty) for subject in YEAR_2_SUBJECTS for _ in range(subject.numbers.get(0, 0))]
        year_2_practicals = [Class(subject, 1, subject.faculty) for subject in YEAR_2_SUBJECTS for _ in range(subject.numbers.get(1, 0))]
        classes = year_2_lectures + year_2_practicals
        mapping = {i: class_ for i, class_ in enumerate(classes)}
        toolbox = create_toolbox(classes, mapping)
        schedule = generate_timetable(classes, mapping, toolbox)  # Correctly call the function
        print_timetable(schedule)
        write_timetable_to_excel(f"2nd_Year_Division_{schedule_num + 1}.xlsx", schedule)

if __name__ == "__main__":
    main()